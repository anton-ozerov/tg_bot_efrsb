import os
from datetime import datetime
from pathlib import Path

import aiohttp
from aiogram import Bot
from openpyxl import load_workbook

from app.data.config import OSINT_1_TOKEN
from app.database.requests import Database
from app.utils.generate_xlsx import generate_new_xlsx


async def fetch_data_dyxless(query: str, token: str = OSINT_1_TOKEN):
    url = "https://api-dyxless.cfd/query"

    payload = {
        "query": query,
        "token": token
    }

    headers = {
        "Content-Type": "application/json"
    }
    async with aiohttp.ClientSession() as session:
        try:
            async with session.post(url, headers=headers, json=payload) as response:
                response.raise_for_status()
                result = await response.json(encoding='utf-8')
            return result
        except aiohttp.ClientError as e:
            print(f"Ошибка запроса: {e}")
            return None


async def update_db(db: Database, revision: int, phones: list[str], emails: list[str]):
    for phone in phones:
        await db.add_phone_to_delo(delo_id=revision, phone_number=phone)
    for email in emails:
        await db.add_email_to_delo(delo_id=revision, email_address=email)


async def add_info(db: Database, file_id: str, bot: Bot):
    file = await bot.get_file(file_id)
    file_path = file.file_path
    output_dir = Path("temp_files")
    output_dir.mkdir(exist_ok=True)
    file_name = output_dir / f"{file_id}.xlsx"
    await bot.download_file(file_path, file_name)

    wb = load_workbook(file_name)
    ws = wb.active  # первая (активная) страница

    # Ищем номер столбца с заголовком "Дата публикации" в первой строке
    date_col = None
    fio_col, inn_col, revision_col = None, None, None
    for cell in ws[1]:  # первая строка
        if cell.value == "Дата публикации":
            date_col = cell.column  # openpyxl 3.0+ — column возвращает номер столбца (int)
        elif cell.value == 'ИНН':
            inn_col = cell.column
        elif cell.value == 'Ревизия':
            revision_col = cell.column

    # Из второй строки в этом столбце берем дату (date_1)
    date_1 = ws.cell(row=2, column=date_col).value

    # Из последней строки в этом столбце берем дату (date_2)
    last_row = ws.max_row
    date_2 = ws.cell(row=last_row, column=date_col).value

    date_1, date_2 = datetime.strptime(date_1, '%Y-%m-%d %H:%M:%S'), datetime.strptime(date_2, '%Y-%m-%d %H:%M:%S')

    for row in range(2, last_row + 1):
        data = f"{ws.cell(row=row, column=inn_col).value}"
        revision = int(ws.cell(row=row, column=revision_col).value)
        res = await fetch_data_dyxless(query=data)
        phones = set()
        emails = set()
        if res is not None:
            for person_info in res['data']:
                if 'number' in person_info:
                    try:
                        int(person_info['number'].replace('+', ''))
                        phones.add(person_info['number'].replace('+', ''))
                    except ValueError:
                        pass
                if 'mail' in person_info and '@' in person_info['mail']:
                    emails.add(person_info['mail'])
                if 'other' in person_info:
                    if '@' in person_info['other']:
                        emails.add(person_info['other'])
                    elif len(person_info['other'].replace('+', '')) == 11:
                        try:
                            int(person_info['other'].replace('+', ''))
                            phones.add(person_info['other'].replace('+', ''))
                        except ValueError:
                            pass
                if 'null' in person_info:
                    for i in person_info['null']:
                        if '@' in i:
                            emails.add(i)
                        elif len(i.replace('+', '')) == 11:
                            try:
                                int(i.replace('+', ''))
                                phones.add(i.replace('+', ''))
                            except ValueError:
                                pass
            await update_db(db=db, revision=revision, phones=list(phones), emails=list(emails))
        else:
            if os.path.exists(file_name):
                os.remove(file_name)
            return None

    if os.path.exists(file_name):
        os.remove(file_name)
    file_name = await generate_new_xlsx(db=db, date_1=date_1, date_2=date_2, is_with_osint=True)
    return file_name
